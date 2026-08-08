"""Export the human-reviewed gap register to CSV and XLSX.

Exports read persisted rows, so they reflect the reviewed state. Any row still in
'AI-proposed' status is explicitly flagged 'UNREVIEWED' in a dedicated column and,
in XLSX, tinted, so a downstream reader can never mistake an un-validated AI
proposal for a signed-off finding.
"""
from __future__ import annotations

import csv
import io
from typing import List

from .models import GapRow, ReviewStatus

COLUMNS = [
    "requirement_id", "requirement_ref", "requirement_title", "requirement_summary",
    "mapped_control_ids", "coverage_status", "confidence", "confidence_band",
    "severity", "review_status", "unreviewed_flag", "assessor",
    "rationale", "cited_control_quote", "recommended_remediation", "reviewer_note",
]


def _row_to_record(row: GapRow) -> dict:
    control_quotes = " | ".join(
        f"[{c.control_id}] \"{c.control_quote}\""
        for c in row.citations if c.control_id and c.control_quote
    )
    return {
        "requirement_id": row.requirement_id,
        "requirement_ref": row.requirement_ref,
        "requirement_title": row.requirement_title,
        "requirement_summary": row.requirement_summary,
        "mapped_control_ids": ", ".join(row.mapped_control_ids) if row.mapped_control_ids else "(none)",
        "coverage_status": row.coverage_status.value,
        "confidence": row.confidence,
        "confidence_band": row.confidence_band.value,
        "severity": row.severity.value,
        "review_status": row.review_status.value,
        "unreviewed_flag": "UNREVIEWED" if row.review_status == ReviewStatus.AI_PROPOSED else "",
        "assessor": row.assessor,
        "rationale": row.rationale,
        "cited_control_quote": control_quotes,
        "recommended_remediation": row.recommended_remediation,
        "reviewer_note": row.reviewer_note,
    }


def to_csv(rows: List[GapRow]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=COLUMNS, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(_row_to_record(row))
    return buf.getvalue().encode("utf-8")


def to_xlsx(rows: List[GapRow], meta: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # --- Register sheet ---
    ws = wb.active
    ws.title = "Gap Register"

    header_fill = PatternFill("solid", fgColor="1F2933")
    header_font = Font(color="FFFFFF", bold=True, name="Calibri")
    unreviewed_fill = PatternFill("solid", fgColor="FFF3CD")
    gap_fill = PatternFill("solid", fgColor="F8D7DA")
    partial_fill = PatternFill("solid", fgColor="FDEBD0")

    for col_idx, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.fill = header_fill
        cell.font = header_font

    for r_idx, row in enumerate(rows, start=2):
        rec = _row_to_record(row)
        for c_idx, col in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=rec[col])
            cell.alignment = Alignment(vertical="top", wrap_text=col in (
                "requirement_summary", "rationale", "cited_control_quote",
                "recommended_remediation", "reviewer_note"))
        if row.review_status == ReviewStatus.AI_PROPOSED:
            ws.cell(row=r_idx, column=COLUMNS.index("unreviewed_flag") + 1).fill = unreviewed_fill
        status_cell = ws.cell(row=r_idx, column=COLUMNS.index("coverage_status") + 1)
        if row.coverage_status.value == "Gap":
            status_cell.fill = gap_fill
        elif row.coverage_status.value == "Partial":
            status_cell.fill = partial_fill

    widths = {
        "requirement_summary": 40, "rationale": 55, "cited_control_quote": 45,
        "recommended_remediation": 45, "reviewer_note": 30, "requirement_title": 24,
    }
    for c_idx, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = widths.get(col, 16)
    ws.freeze_panes = "A2"

    # --- Provenance sheet ---
    ws2 = wb.create_sheet("Run Info")
    prov = [
        ("ControlGap export", ""),
        ("DISCLAIMER", "Decision-support only. Not legal or compliance advice. Control data is SYNTHETIC. Outputs require human validation."),
        ("Standard", str(meta.get("standard", {}).get("standard_name", ""))),
        ("Assessor", str(meta.get("assessor", {}).get("assessor", ""))),
        ("Assessor is stub", str(meta.get("assessor", {}).get("is_stub", ""))),
        ("Embedder backend", str(meta.get("embedder", {}).get("backend", ""))),
        ("Embedder is fallback", str(meta.get("embedder", {}).get("is_fallback", ""))),
        ("Top-k", str(meta.get("top_k", ""))),
        ("Total rows", str(len(rows))),
        ("Unreviewed rows", str(sum(1 for r in rows if r.review_status == ReviewStatus.AI_PROPOSED))),
    ]
    for i, (k, v) in enumerate(prov, start=1):
        ws2.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=v).alignment = Alignment(wrap_text=True, vertical="top")
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 90

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
